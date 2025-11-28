import codecs
from fastapi import APIRouter, Query, HTTPException, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from backend_api.database import get_db
from typing import List, Optional
import io
import csv
from sqlalchemy import text
from datetime import datetime
from pydantic import BaseModel

# 新增依赖
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment

router = APIRouter(prefix="/api/stock/history", tags=["StockHistory"])

# 计算5天升跌请求模型
class CalculateFiveDayChangeRequest(BaseModel):
    stock_code: str
    start_date: str
    end_date: str
    extended_end_date: Optional[str] = None  # 扩展后的结束日期（用于60天涨跌计算）

def format_date_yyyymmdd(date_str: Optional[str]) -> Optional[str]:
    if not date_str:
        return None
    # 支持 "YYYY-MM-DD"、"YYYY/MM/DD"、"YYYY.MM.DD" 等
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
        except Exception:
            continue
    # 如果本身就是8位数字，尝试转为YYYY-MM-DD格式
    if len(date_str) == 8 and date_str.isdigit():
        try:
            return datetime.strptime(date_str, "%Y%m%d").strftime("%Y-%m-%d")
        except Exception:
            pass
    return date_str  # fallback

@router.get("")
def get_stock_history(
    code: str = Query(...),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    include_notes: bool = Query(True, description="是否包含交易备注"),
    db: Session = Depends(get_db)
):
    start_date_fmt = format_date_yyyymmdd(start_date)
    end_date_fmt = format_date_yyyymmdd(end_date)
    print(f"[get_stock_history] 输入参数: code={code}, start_date={start_date_fmt}, end_date={end_date_fmt}, page={page}, size={size}, include_notes={include_notes}")
    
    # 1. 先尝试从A股历史行情表查询
    items = []
    total = 0
    is_hk_data = False
    
    if include_notes:
        # 使用视图查询，包含交易备注
        query_a = """
            SELECT 
                h.code, h.name, h.date, h.open, h.close, h.high, h.low, 
                h.volume, h.amount, h.change_percent, h.change, h.turnover_rate,
                h.cumulative_change_percent, h.five_day_change_percent, h.ten_day_change_percent, h.thirty_day_change_percent, h.sixty_day_change_percent, h.remarks,
                COALESCE(tn.notes, '') as user_notes,
                COALESCE(tn.strategy_type, '') as strategy_type,
                COALESCE(tn.risk_level, '') as risk_level,
                COALESCE(tn.created_by, '') as notes_creator,
                tn.created_at as notes_created_at,
                tn.updated_at as notes_updated_at
            FROM historical_quotes h
            LEFT JOIN trading_notes tn ON h.code = tn.stock_code AND h.date::date = tn.trade_date
            WHERE h.code = :code
        """
    else:
        # 只查询基础历史行情数据
        query_a = """
            SELECT 
                code, name, date, open, close, high, low, 
                volume, amount, change_percent, change, turnover_rate,
                cumulative_change_percent, five_day_change_percent, ten_day_change_percent, thirty_day_change_percent, sixty_day_change_percent, remarks
            FROM historical_quotes 
            WHERE code = :code
        """
    
    params_a = {"code": code}
    if start_date_fmt:
        query_a += " AND date >= :start_date"
        params_a["start_date"] = start_date_fmt
    if end_date_fmt:
        query_a += " AND date <= :end_date"
        params_a["end_date"] = end_date_fmt
    query_a += " ORDER BY date DESC"
    
    try:
        # 先检查A股表是否有数据
        count_query_a = f"SELECT COUNT(*) FROM ({query_a})"
        total_a = db.execute(text(count_query_a), params_a).scalar()
        
        if total_a > 0:
            # A股有数据，使用A股数据
            total = total_a
            query_a += " LIMIT :limit OFFSET :offset"
            params_a["limit"] = str(size)
            params_a["offset"] = str((page - 1) * size)
            result = db.execute(text(query_a), params_a)
            
            for row in result.fetchall():
                item = {
                    "code": row[0],
                    "name": row[1],
                    "date": row[2],
                    "open": row[3],
                    "close": row[4],
                    "high": row[5],
                    "low": row[6],
                    "volume": row[7],
                    "amount": row[8],
                    "change_percent": row[9],
                    "change": row[10],
                    "turnover_rate": row[11],
                    "cumulative_change_percent": row[12],
                    "five_day_change_percent": row[13],
                    "ten_day_change_percent": row[14],
                    "thirty_day_change_percent": row[15],
                    "sixty_day_change_percent": row[16],
                    "remarks": row[17] if len(row) > 17 else None
                }
                
                # 如果包含备注，添加备注相关字段
                if include_notes and len(row) > 18:
                    item.update({
                        "user_notes": row[18],
                        "strategy_type": row[19],
                        "risk_level": row[20],
                        "notes_creator": row[21],
                        "notes_created_at": row[22],
                        "notes_updated_at": row[23]
                    })
                
                items.append(item)
        else:
            # A股表完全没有数据，才从港股历史行情表查询
            print(f"[get_stock_history] A股表无数据，从港股表查询: code={code}")
            is_hk_data = True
            
            if include_notes:
                query_hk = """
                    SELECT 
                        h.code, h.name, h.date, h.open, h.close, h.high, h.low, 
                        h.volume, h.amount, h.change_percent, h.change_amount, h.turnover_rate,
                        NULL as cumulative_change_percent, h.five_day_change_percent, h.ten_day_change_percent, h.thirty_day_change_percent, h.sixty_day_change_percent, NULL as remarks,
                        COALESCE(tn.notes, '') as user_notes,
                        COALESCE(tn.strategy_type, '') as strategy_type,
                        COALESCE(tn.risk_level, '') as risk_level,
                        COALESCE(tn.created_by, '') as notes_creator,
                        tn.created_at as notes_created_at,
                        tn.updated_at as notes_updated_at
                    FROM historical_quotes_hk h
                    LEFT JOIN trading_notes tn ON h.code = tn.stock_code AND CAST(h.date AS DATE) = CAST(tn.trade_date AS DATE)
                    WHERE h.code = :code
                """
            else:
                query_hk = """
                    SELECT 
                        code, name, date, open, close, high, low, 
                        volume, amount, change_percent, change_amount, turnover_rate,
                        NULL as cumulative_change_percent, five_day_change_percent, ten_day_change_percent, thirty_day_change_percent, sixty_day_change_percent, NULL as remarks
                    FROM historical_quotes_hk 
                    WHERE code = :code
                """
            
            params_hk = {"code": code}
            if start_date_fmt:
                query_hk += " AND date >= :start_date"
                params_hk["start_date"] = start_date_fmt
            if end_date_fmt:
                query_hk += " AND date <= :end_date"
                params_hk["end_date"] = end_date_fmt
            query_hk += " ORDER BY date DESC"
            
            count_query_hk = f"SELECT COUNT(*) FROM ({query_hk})"
            total = db.execute(text(count_query_hk), params_hk).scalar()
            
            if total > 0:
                query_hk += " LIMIT :limit OFFSET :offset"
                params_hk["limit"] = str(size)
                params_hk["offset"] = str((page - 1) * size)
                result = db.execute(text(query_hk), params_hk)
                
                for row in result.fetchall():
                    item = {
                        "code": row[0],
                        "name": row[1],
                        "date": row[2],
                        "open": row[3],
                        "close": row[4],
                        "high": row[5],
                        "low": row[6],
                        "volume": row[7],
                        "amount": row[8],
                        "change_percent": row[9],
                        "change": row[10],  # 港股使用change_amount，映射为change
                        "turnover_rate": row[11],
                        "cumulative_change_percent": row[12],  # 港股可能为NULL
                        "five_day_change_percent": row[13],
                        "ten_day_change_percent": row[14],
                        "thirty_day_change_percent": row[15],
                        "sixty_day_change_percent": row[16],
                        "remarks": row[17] if len(row) > 17 else None
                    }
                    
                    # 如果包含备注，添加备注相关字段
                    if include_notes and len(row) > 18:
                        item.update({
                            "user_notes": row[18],
                            "strategy_type": row[19],
                            "risk_level": row[20],
                            "notes_creator": row[21],
                            "notes_created_at": row[22],
                            "notes_updated_at": row[23]
                        })
                    
                    items.append(item)
    except Exception as e:
        print(f"[get_stock_history] 查询异常: {e}")
        import traceback
        traceback.print_exc()
        return {"items": [], "total": 0}

    # 遍历items，如果换手率为None，则用akshare查询并回写数据库（支持A股和港股）
    try:
        import akshare as ak
    except ImportError:
        ak = None  # 如果akshare未安装，后面查询会报错

    # 先判断是否存在换手率为None或0的记录
    need_fill_items = [item for item in items if item.get("turnover_rate") is None or item.get("turnover_rate") == 0]
    if ak and need_fill_items and start_date and end_date:
        # 统一从前台参数获取start_date与end_date，批量查询
        # start_date, end_date 已在上文Query参数中拿到
        try:
            # akshare日期要求为yyyymmdd字符串
            code = need_fill_items[0]["code"]
            
            if is_hk_data:
                # 港股：使用stock_hk_hist接口
                hist_df = ak.stock_hk_hist(symbol=code, period='daily', start_date=start_date.replace("-", ""), end_date=end_date.replace("-", ""), adjust="")
                print(f"[港股] akshare查询结果: {hist_df}")
                table_name = "historical_quotes_hk"
            else:
                # A股：使用stock_zh_a_hist接口
                hist_df = ak.stock_zh_a_hist(symbol=code, start_date=start_date.replace("-", ""), end_date=end_date.replace("-", ""), adjust="")
                print(f"[A股] akshare查询结果: {hist_df}")
                table_name = "historical_quotes"
            
            # 将日期设为index，便于查找
            if not hist_df.empty and "换手率" in hist_df.columns and "日期" in hist_df.columns:
                # 处理日期格式：akshare返回的日期可能是YYYYMMDD或YYYY-MM-DD格式
                hist_df["日期"] = hist_df["日期"].astype(str)
                # 统一转换为YYYY-MM-DD格式
                date_formatted = []
                for d in hist_df["日期"]:
                    if len(d) == 8 and d.isdigit():
                        date_formatted.append(f"{d[:4]}-{d[4:6]}-{d[6:8]}")
                    else:
                        date_formatted.append(d)
                hist_df["日期"] = date_formatted
                hist_df.set_index("日期", inplace=True)
                
                for item in need_fill_items:
                    code = item["code"]
                    date = item["date"]  # 格式已为yyyy-mm-dd
                    q_date = date  # 结果集日期已经是yyyy-mm-dd, 无需格式化
                    turnover = None
                    try:
                        # 从结果集中查找对应日期
                        if q_date in hist_df.index and not pd.isna(hist_df.loc[q_date, "换手率"]):
                            val = hist_df.loc[q_date, "换手率"]
                            if isinstance(val, str):
                                val = val.replace("%", "")
                            turnover = float(val)
                            item["turnover_rate"] = turnover
                            # 回写数据库
                            db.execute(
                                text(
                                    f"UPDATE {table_name} SET turnover_rate = :turnover_rate WHERE code = :code AND date = :date"
                                ),
                                {"turnover_rate": turnover, "code": code, "date": date}
                            )
                            db.commit()
                            # 记录换手率更新的成功次数
                            if not hasattr(get_stock_history, "_turnover_update_count"):
                                get_stock_history._turnover_update_count = 0
                            get_stock_history._turnover_update_count += 1
                            print(f"成功更新【换手率】次数：{get_stock_history._turnover_update_count}, code={code}, date={date}, turnover={turnover}")
                    except Exception as ex:
                        print(f"通过akshare批量补充换手率失败: code={code}, date={date}, error={str(ex)}")
        except Exception as e:
            print(f"调用akshare接口批量获取数据异常: {e}")
            import traceback
            traceback.print_exc()

    print(f"[get_stock_history] 输出: total={total}, items_count={len(items)}")
    return {"items": items, "total": total}

@router.get("/export")
def export_stock_history(
    code: str = Query(...),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    include_notes: bool = Query(True, description="是否包含交易备注"),
    format: str = Query("excel", description="导出格式: csv 或 excel"),
    db: Session = Depends(get_db)
):
    start_date_fmt = format_date_yyyymmdd(start_date)
    end_date_fmt = format_date_yyyymmdd(end_date)
    print(f"[export_stock_history] 输入参数: code={code}, start_date={start_date_fmt}, end_date={end_date_fmt}, include_notes={include_notes}")
    
    # 1. 先尝试从A股历史行情表查询
    rows = []
    is_hk_data = False
    
    if include_notes:
        # 包含备注的查询，关联trading_notes表
        base_query_a = """
            SELECT 
                hq.code, hq.name, hq.date, hq.open, hq.close, hq.high, hq.low, 
                hq.volume, hq.amount, hq.change_percent, hq.change, hq.turnover_rate,
                COALESCE(hq.cumulative_change_percent, 0) as cumulative_change_percent, 
                COALESCE(hq.five_day_change_percent, 0) as five_day_change_percent,
                COALESCE(hq.ten_day_change_percent, 0) as ten_day_change_percent,
                COALESCE(hq.thirty_day_change_percent, 0) as thirty_day_change_percent,
                COALESCE(hq.sixty_day_change_percent, 0) as sixty_day_change_percent,
                COALESCE(tn.notes, '') as user_notes,
                COALESCE(tn.strategy_type, '') as strategy_type,
                COALESCE(tn.risk_level, '') as risk_level
            FROM historical_quotes hq
            LEFT JOIN trading_notes tn ON hq.code = tn.stock_code AND CAST(hq.date AS DATE) = CAST(tn.trade_date AS DATE)
            WHERE hq.code = :code
        """
    else:
        # 不包含备注的查询
        base_query_a = """
            SELECT 
                code, name, date, open, close, high, low, 
                volume, amount, change_percent, change, turnover_rate,
                COALESCE(cumulative_change_percent, 0) as cumulative_change_percent, 
                COALESCE(five_day_change_percent, 0) as five_day_change_percent,
                COALESCE(ten_day_change_percent, 0) as ten_day_change_percent,
                COALESCE(thirty_day_change_percent, 0) as thirty_day_change_percent,
                COALESCE(sixty_day_change_percent, 0) as sixty_day_change_percent,
                '' as user_notes,
                '' as strategy_type,
                '' as risk_level
            FROM historical_quotes 
            WHERE code = :code
        """
    
    params_a = {"code": code}
    if start_date_fmt:
        base_query_a += " AND hq.date >= :start_date" if include_notes else " AND date >= :start_date"
        params_a["start_date"] = start_date_fmt
    if end_date_fmt:
        base_query_a += " AND hq.date <= :end_date" if include_notes else " AND date <= :end_date"
        params_a["end_date"] = end_date_fmt
    base_query_a += " ORDER BY hq.date DESC" if include_notes else " ORDER BY date DESC"
    
    try:
        result_a = db.execute(text(base_query_a), params_a)
        rows = result_a.fetchall()
        
        if not rows:
            # A股表完全没有数据，才从港股历史行情表查询
            print(f"[export_stock_history] A股表无数据，从港股表查询: code={code}")
            is_hk_data = True
            
            if include_notes:
                base_query_hk = """
                    SELECT 
                        hq.code, hq.name, hq.date, hq.open, hq.close, hq.high, hq.low, 
                        hq.volume, hq.amount, hq.change_percent, hq.change_amount as change, hq.turnover_rate,
                        0 as cumulative_change_percent, 
                        COALESCE(hq.five_day_change_percent, 0) as five_day_change_percent,
                        COALESCE(hq.ten_day_change_percent, 0) as ten_day_change_percent,
                        COALESCE(hq.thirty_day_change_percent, 0) as thirty_day_change_percent,
                        COALESCE(hq.sixty_day_change_percent, 0) as sixty_day_change_percent,
                        COALESCE(tn.notes, '') as user_notes,
                        COALESCE(tn.strategy_type, '') as strategy_type,
                        COALESCE(tn.risk_level, '') as risk_level
                    FROM historical_quotes_hk hq
                    LEFT JOIN trading_notes tn ON hq.code = tn.stock_code AND CAST(hq.date AS DATE) = CAST(tn.trade_date AS DATE)
                    WHERE hq.code = :code
                """
            else:
                base_query_hk = """
                    SELECT 
                        code, name, date, open, close, high, low, 
                        volume, amount, change_percent, change_amount as change, turnover_rate,
                        0 as cumulative_change_percent, 
                        COALESCE(five_day_change_percent, 0) as five_day_change_percent,
                        COALESCE(ten_day_change_percent, 0) as ten_day_change_percent,
                        COALESCE(thirty_day_change_percent, 0) as thirty_day_change_percent,
                        COALESCE(sixty_day_change_percent, 0) as sixty_day_change_percent,
                        '' as user_notes,
                        '' as strategy_type,
                        '' as risk_level
                    FROM historical_quotes_hk 
                    WHERE code = :code
                """
            
            params_hk = {"code": code}
            if start_date_fmt:
                base_query_hk += " AND hq.date >= :start_date" if include_notes else " AND date >= :start_date"
                params_hk["start_date"] = start_date_fmt
            if end_date_fmt:
                base_query_hk += " AND hq.date <= :end_date" if include_notes else " AND date <= :end_date"
                params_hk["end_date"] = end_date_fmt
            base_query_hk += " ORDER BY hq.date DESC" if include_notes else " ORDER BY date DESC"
            
            result_hk = db.execute(text(base_query_hk), params_hk)
            rows = result_hk.fetchall()
            
            if not rows:
                raise HTTPException(status_code=404, detail="未找到数据")
    except HTTPException:
        raise
    except Exception as e:
        print(f"[export_stock_history] 查询异常: {e}")
        raise HTTPException(status_code=500, detail=f"查询数据失败: {str(e)}")
    
    # 检查是否有备注数据
    has_notes_data = False
    if include_notes:
        try:
            # 首先检查trading_notes表是否存在
            table_check_query = """
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_schema = 'public' 
                    AND table_name = 'trading_notes'
                );
            """
            table_exists = db.execute(text(table_check_query)).scalar()
            
            if not table_exists:
                print(f"[export_stock_history] trading_notes表不存在，跳过备注查询")
                has_notes_data = False
            else:
                # 检查是否有任何备注数据
                notes_check_query = """
                    SELECT COUNT(*) FROM trading_notes 
                    WHERE stock_code = :code 
                    AND (notes IS NOT NULL AND notes != '')
                """
                notes_count = db.execute(text(notes_check_query), {"code": code}).scalar()
                has_notes_data = notes_count > 0
                print(f"[export_stock_history] 备注数据检查: 找到 {notes_count} 条有备注的记录")
                
        except Exception as e:
            print(f"[export_stock_history] 检查备注数据时出错: {e}")
            has_notes_data = False
    
    # 根据格式选择导出方式
    fmt = format.lower()
    if fmt == "excel":
        return export_to_excel(rows, include_notes, has_notes_data, code)
    if fmt == "text":
        return export_to_text(rows, include_notes, has_notes_data, code)
    if fmt == "csv":
        return export_to_csv(rows, include_notes, has_notes_data, code)
    raise HTTPException(status_code=400, detail=f"不支持的导出格式: {format}")

def prepare_export_rows(rows, include_notes, has_notes_data):
    """根据导出配置格式化行数据"""
    # 添加数据格式化函数
    def format_volume(volume):
        """格式化成交量：超过亿以亿为单位，其他值以万为单位，2位小数，四舍五入"""
        if volume is None:
            return '-'
        vol = float(volume)
        # 1亿 = 10000万
        if vol >= 100000000:  # 超过1亿，以亿为单位
            vol_yi = round(vol / 100000000, 2)
            return f"{vol_yi:.2f}亿"
        else:  # 其他值以万为单位
            vol_wan = round(vol / 10000, 2)
            return f"{vol_wan:.2f}万"
    
    def format_amount(amount):
        """格式化成交额：超过亿以亿为单位，其他值以万为单位，2位小数，四舍五入"""
        if amount is None:
            return '-'
        amt = float(amount)
        # 1亿 = 10000万
        if amt >= 100000000:  # 超过1亿，以亿为单位
            amt_yi = round(amt / 100000000, 2)
            return f"{amt_yi:.2f}亿"
        else:  # 其他值以万为单位
            amt_wan = round(amt / 10000, 2)
            return f"{amt_wan:.2f}万"
    
    def format_percent(value):
        """格式化百分比，小数点后2位，四舍五入"""
        if value is None:
            return '-'
        # 保留2位小数，四舍五入
        pct = round(float(value), 2)
        return f"{pct:.2f}%"
    
    def format_price(value):
        """格式化价格"""
        if value is None:
            return '-'
        return f"{float(value):.2f}"
    
    # 根据最新列表字段顺序：股票名称、日期、开盘、收盘、涨跌额、成交额、涨跌幅、成交量、换手率、5天涨跌%、10天涨跌%、30天涨跌%、60天涨跌%、备注
    if include_notes and has_notes_data:
        headers = [
            "股票名称", "日期", "开盘", "收盘", "涨跌额", "成交额", "涨跌幅", 
            "成交量", "换手率", "5天涨跌%", "10天涨跌%", "30天涨跌%", "60天涨跌%", "备注"
        ]
    else:
        headers = [
            "股票名称", "日期", "开盘", "收盘", "涨跌额", "成交额", "涨跌幅", 
            "成交量", "换手率", "5天涨跌%", "10天涨跌%", "30天涨跌%", "60天涨跌%", "备注"
        ]
    
    data_rows = []
    for row in rows:
        # row结构: code, name, date, open, close, high, low, volume, amount, change_percent, change, turnover_rate,
        #          cumulative_change_percent, five_day_change_percent, ten_day_change_percent, thirty_day_change_percent, sixty_day_change_percent,
        #          user_notes, strategy_type, risk_level
        # 按照新顺序：股票名称、日期、开盘、收盘、涨跌额、成交额、涨跌幅、成交量、换手率、5天涨跌%、10天涨跌%、30天涨跌%、60天涨跌%、备注
        if include_notes and has_notes_data:
            # 合并备注信息
            notes_text = row[17] if row[17] else ''
            if row[18]:  # strategy_type
                notes_text += f" [策略:{row[18]}]" if notes_text else f"[策略:{row[18]}]"
            if row[19]:  # risk_level
                notes_text += f" [风险:{row[19]}]" if notes_text else f"[风险:{row[19]}]"
            
            data_rows.append([
                row[1],  # 股票名称
                row[2],  # 日期
                format_price(row[3]),  # 开盘
                format_price(row[4]),  # 收盘
                format_price(row[10]),  # 涨跌额
                format_amount(row[8]),  # 成交额
                format_percent(row[9]),  # 涨跌幅
                format_volume(row[7]),  # 成交量
                format_percent(row[11]),  # 换手率
                format_percent(row[13]),  # 5天涨跌%
                format_percent(row[14]),  # 10天涨跌%
                format_percent(row[15]),  # 30天涨跌%
                format_percent(row[16]),  # 60天涨跌%
                notes_text  # 备注
            ])
        else:
            data_rows.append([
                row[1],  # 股票名称
                row[2],  # 日期
                format_price(row[3]),  # 开盘
                format_price(row[4]),  # 收盘
                format_price(row[10]),  # 涨跌额
                format_amount(row[8]),  # 成交额
                format_percent(row[9]),  # 涨跌幅
                format_volume(row[7]),  # 成交量
                format_percent(row[11]),  # 换手率
                format_percent(row[13]),  # 5天涨跌%
                format_percent(row[14]),  # 10天涨跌%
                format_percent(row[15]),  # 30天涨跌%
                format_percent(row[16]),  # 60天涨跌%
                row[17] if row[17] else ''  # 备注
            ])

    return headers, data_rows

def export_to_csv(rows, include_notes, has_notes_data, code):
    """导出为CSV格式"""
    headers, data_rows = prepare_export_rows(rows, include_notes, has_notes_data)
    
    # 创建CSV内容
    output = io.StringIO()
    
    # 添加BOM头，解决Excel打开中文乱码问题
    output.write('\ufeff')
    
    writer = csv.writer(output)
    
    # 根据是否有备注数据设置不同的CSV头
    writer.writerow(headers)
    
    # 写入数据
    writer.writerows(data_rows)
    
    output.seek(0)
    
    # 生成文件名
    filename = f"{code}_historical_quotes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

def export_to_text(rows, include_notes, has_notes_data, code):
    """导出为制表符分隔的文本格式"""
    headers, data_rows = prepare_export_rows(rows, include_notes, has_notes_data)

    output = io.StringIO()
    # 添加BOM，确保 Windows 记事本等工具显示正常
    output.write('\ufeff')
    output.write('\t'.join(headers) + '\n')

    for row in data_rows:
        output.write('\t'.join(str(value) for value in row) + '\n')

    output.seek(0)
    filename = f"{code}_historical_quotes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

def export_to_excel(rows, include_notes, has_notes_data, code):
    """导出为Excel格式，支持颜色效果"""
    # 数据格式化函数
    def format_volume_text(volume):
        """格式化成交量为文本：超过亿以亿为单位，其他值以万为单位，2位小数，四舍五入"""
        if volume is None:
            return None
        vol = float(volume)
        # 1亿 = 10000万
        if vol >= 100000000:  # 超过1亿，以亿为单位
            vol_yi = round(vol / 100000000, 2)
            return f"{vol_yi:.2f}亿"
        else:  # 其他值以万为单位
            vol_wan = round(vol / 10000, 2)
            return f"{vol_wan:.2f}万"
    
    def format_amount_text(amount):
        """格式化成交额为文本：超过亿以亿为单位，其他值以万为单位，2位小数，四舍五入"""
        if amount is None:
            return None
        amt = float(amount)
        # 1亿 = 10000万
        if amt >= 100000000:  # 超过1亿，以亿为单位
            amt_yi = round(amt / 100000000, 2)
            return f"{amt_yi:.2f}亿"
        else:  # 其他值以万为单位
            amt_wan = round(amt / 10000, 2)
            return f"{amt_wan:.2f}万"
    
    def format_percent_text(value):
        """格式化百分比为文本：如0.45%，小数点后2位，四舍五入"""
        if value is None:
            return None
        pct = round(float(value), 2)
        return f"{pct:.2f}%"
    
    def safe_float(value):
        """安全转换为浮点数"""
        if value is None:
            return None
        return float(value)
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active  # type: ignore
    assert ws is not None
    ws.title = "历史行情数据"
    
    # 定义颜色样式
    red_font = Font(color="FF0000", bold=True)    # 上涨红色
    green_font = Font(color="00AA00", bold=True)  # 下跌绿色
    normal_font = Font(color="000000")
    header_font = Font(bold=True)
    
    # 设置表头 - 按照最新列表字段顺序：股票名称、日期、开盘、收盘、涨跌额、成交额、涨跌幅、成交量、换手率、5天涨跌%、10天涨跌%、30天涨跌%、60天涨跌%、备注
    headers = [
        "股票名称", "日期", "开盘", "收盘", "涨跌额", "成交额", "涨跌幅", 
        "成交量", "换手率", "5天涨跌%", "10天涨跌%", "30天涨跌%", "60天涨跌%", "备注"
    ]
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
    
    # 写入数据并应用颜色格式
    # row结构: code, name, date, open, close, high, low, volume, amount, change_percent, change, turnover_rate,
    #          cumulative_change_percent, five_day_change_percent, ten_day_change_percent, thirty_day_change_percent, sixty_day_change_percent,
    #          user_notes, strategy_type, risk_level
    for row_idx, row in enumerate(rows, 2):
        col_idx = 1
        
        # 股票名称
        ws.cell(row=row_idx, column=col_idx, value=row[1])
        col_idx += 1
        
        # 日期
        ws.cell(row=row_idx, column=col_idx, value=row[2])
        col_idx += 1
        
        # 开盘
        ws.cell(row=row_idx, column=col_idx, value=safe_float(row[3]))
        col_idx += 1
        
        # 收盘价（需要颜色格式）
        change_percent = safe_float(row[9])
        close_cell = ws.cell(row=row_idx, column=col_idx, value=safe_float(row[4]))
        if change_percent is not None:
            if change_percent > 0:
                close_cell.font = red_font
            elif change_percent < 0:
                close_cell.font = green_font
        col_idx += 1
        
        # 涨跌额（需要颜色格式）
        change_cell = ws.cell(row=row_idx, column=col_idx, value=safe_float(row[10]))
        if change_percent is not None:
            if change_percent > 0:
                change_cell.font = red_font
            elif change_percent < 0:
                change_cell.font = green_font
        col_idx += 1
        
        # 成交额（需要颜色格式，跟前一天成交额比较），显示为带单位的文本，强制文本格式
        current_amount = safe_float(row[8])
        amount_text = format_amount_text(row[8])
        amount_cell = ws.cell(row=row_idx, column=col_idx, value=amount_text)
        amount_cell.number_format = '@'  # 强制文本格式
        # 获取前一天的成交额（rows是按日期倒序排列的，所以下一行索引是前一天）
        # row_idx从2开始，对应rows[0]（最新日期），rows[1]是前一天
        current_row_index = row_idx - 2  # 转换为rows数组索引
        if current_row_index + 1 < len(rows):  # 确保有前一天的数据
            prev_row = rows[current_row_index + 1]  # 前一天的数据
            prev_amount = safe_float(prev_row[8])
            if current_amount is not None and prev_amount is not None:
                if current_amount > prev_amount:
                    amount_cell.font = red_font  # 比前一天多，红色
                elif current_amount < prev_amount:
                    amount_cell.font = green_font  # 比前一天少，绿色
                # 相等则不变（使用默认字体）
        col_idx += 1
        
        # 涨跌幅%（需要颜色格式），显示格式如0.45%，小数点后2位，四舍五入，强制文本格式
        change_pct_text = format_percent_text(change_percent)
        change_pct_cell = ws.cell(row=row_idx, column=col_idx, value=change_pct_text)
        change_pct_cell.number_format = '@'  # 强制文本格式
        if change_percent is not None:
            if change_percent > 0:
                change_pct_cell.font = red_font
            elif change_percent < 0:
                change_pct_cell.font = green_font
        col_idx += 1
        
        # 成交量（超过亿以亿为单位，其他值以万为单位，小数点后2位，四舍五入），强制文本格式
        volume_text = format_volume_text(row[7])
        volume_cell = ws.cell(row=row_idx, column=col_idx, value=volume_text)
        volume_cell.number_format = '@'  # 强制文本格式
        col_idx += 1
        
        # 换手率%，显示格式如0.45%，小数点后2位，四舍五入，强制文本格式
        turnover_rate_text = format_percent_text(row[11])
        turnover_rate_cell = ws.cell(row=row_idx, column=col_idx, value=turnover_rate_text)
        turnover_rate_cell.number_format = '@'  # 强制文本格式
        col_idx += 1
        
        # 5天涨跌%（需要颜色格式），显示格式如0.45%，小数点后2位，四舍五入，强制文本格式
        five_day_pct = safe_float(row[13])
        five_day_pct_text = format_percent_text(five_day_pct)
        five_day_cell = ws.cell(row=row_idx, column=col_idx, value=five_day_pct_text)
        five_day_cell.number_format = '@'  # 强制文本格式
        if five_day_pct is not None:
            if five_day_pct > 0:
                five_day_cell.font = red_font
            elif five_day_pct < 0:
                five_day_cell.font = green_font
        col_idx += 1
        
        # 10天涨跌%（需要颜色格式），显示格式如0.45%，小数点后2位，四舍五入，强制文本格式
        ten_day_pct = safe_float(row[14])
        ten_day_pct_text = format_percent_text(ten_day_pct)
        ten_day_cell = ws.cell(row=row_idx, column=col_idx, value=ten_day_pct_text)
        ten_day_cell.number_format = '@'  # 强制文本格式
        if ten_day_pct is not None:
            if ten_day_pct > 0:
                ten_day_cell.font = red_font
            elif ten_day_pct < 0:
                ten_day_cell.font = green_font
        col_idx += 1
        
        # 30天涨跌%（需要颜色格式），显示格式如0.45%，小数点后2位，四舍五入，强制文本格式
        thirty_day_pct = safe_float(row[15])
        thirty_day_pct_text = format_percent_text(thirty_day_pct)
        thirty_day_cell = ws.cell(row=row_idx, column=col_idx, value=thirty_day_pct_text)
        thirty_day_cell.number_format = '@'  # 强制文本格式
        if thirty_day_pct is not None:
            if thirty_day_pct > 0:
                thirty_day_cell.font = red_font
            elif thirty_day_pct < 0:
                thirty_day_cell.font = green_font
        col_idx += 1
        
        # 60天涨跌%（需要颜色格式），显示格式如0.45%，小数点后2位，四舍五入，强制文本格式
        sixty_day_pct = safe_float(row[16])
        sixty_day_pct_text = format_percent_text(sixty_day_pct)
        sixty_day_cell = ws.cell(row=row_idx, column=col_idx, value=sixty_day_pct_text)
        sixty_day_cell.number_format = '@'  # 强制文本格式
        if sixty_day_pct is not None:
            if sixty_day_pct > 0:
                sixty_day_cell.font = red_font
            elif sixty_day_pct < 0:
                sixty_day_cell.font = green_font
        col_idx += 1
        
        # 备注
        if include_notes and has_notes_data:
            # 合并备注信息
            notes_text = row[17] if row[17] else ''
            if row[18]:  # strategy_type
                notes_text += f" [策略:{row[18]}]" if notes_text else f"[策略:{row[18]}]"
            if row[19]:  # risk_level
                notes_text += f" [风险:{row[19]}]" if notes_text else f"[风险:{row[19]}]"
            ws.cell(row=row_idx, column=col_idx, value=notes_text)
        else:
            ws.cell(row=row_idx, column=col_idx, value=row[17] if row[17] else '')
        
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 生成文件名
    filename = f"{code}_historical_quotes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )

def _get_date_before_business_days(date_str: str, business_days: int) -> str:
    """
    计算指定日期前N个工作日的日期
    
    Args:
        date_str: 日期字符串 (YYYY-MM-DD)
        business_days: 工作日数量
        
    Returns:
        str: 前N个工作日的日期
    """
    from datetime import datetime, timedelta
    
    current_date = datetime.strptime(date_str, '%Y-%m-%d')
    days_back = 0
    business_days_count = 0
    check_date = current_date  # 初始化变量
    
    while business_days_count < business_days:
        days_back += 1
        check_date = current_date - timedelta(days=days_back)
        
        # 跳过周末（周六=5，周日=6）
        if check_date.weekday() < 5:  # 0-4 表示周一到周五
            business_days_count += 1
    
    return check_date.strftime('%Y-%m-%d')

@router.post("/calculate_five_day_change")
def calculate_five_day_change(
    request: CalculateFiveDayChangeRequest,
    db: Session = Depends(get_db)
):
    """计算指定日期范围内股票的5天升跌%（支持A股和港股）"""
    try:
        # 格式化日期
        start_date_fmt = format_date_yyyymmdd(request.start_date)
        end_date_fmt = format_date_yyyymmdd(request.end_date)
        
        if not start_date_fmt or not end_date_fmt:
            raise HTTPException(status_code=400, detail="日期格式无效")
        
        # 如果提供了扩展后的结束日期，使用它来扩展查询和更新范围
        query_end_date = format_date_yyyymmdd(request.extended_end_date) if request.extended_end_date else end_date_fmt
        
        print(f"[calculate_five_day_change] 开始计算股票 {request.stock_code} 在 {start_date_fmt} 到 {end_date_fmt} 期间的5天升跌%（查询范围扩展到 {query_end_date}）")
        
        # 为了确保最后5条记录也能计算5天升跌%，需要查询更早的数据
        # 计算开始日期前5个工作日的数据
        extended_start_date = _get_date_before_business_days(start_date_fmt, 5)
        
        # 1. 先尝试从A股表查询
        extended_query_a = """
            SELECT date, close
            FROM historical_quotes 
            WHERE code = :stock_code 
            AND date >= :extended_start_date 
            AND date <= :end_date
            ORDER BY date ASC
        """
        
        result_a = db.execute(text(extended_query_a), {
            "stock_code": request.stock_code,
            "extended_start_date": extended_start_date,
            "end_date": query_end_date
        })
        
        quotes = result_a.fetchall()
        is_hk_data = False
        
        # 2. 如果A股表完全没有数据，才从港股表查询
        if len(quotes) == 0:
            print(f"[calculate_five_day_change] A股表无数据，从港股表查询: code={request.stock_code}")
            extended_query_hk = """
                SELECT date, close
                FROM historical_quotes_hk 
                WHERE code = :stock_code 
                AND date >= :extended_start_date 
                AND date <= :end_date
                ORDER BY date ASC
            """
            
            result_hk = db.execute(text(extended_query_hk), {
                "stock_code": request.stock_code,
                "extended_start_date": extended_start_date,
                "end_date": query_end_date
            })
            
            quotes = result_hk.fetchall()
            is_hk_data = True
        
        if len(quotes) < 6:
            raise HTTPException(status_code=400, detail="数据不足6天，无法计算5天升跌%")
        
        # 计算5天升跌%
        updated_count = 0
        table_name = "historical_quotes_hk" if is_hk_data else "historical_quotes"
        
        for i in range(5, len(quotes)):
            current_quote = quotes[i]
            prev_quote = quotes[i-5]  # 5天前的数据
            
            # 更新到扩展结束日期范围内的所有记录，确保最后5天的数据也能被计算
            if current_quote.date >= start_date_fmt and current_quote.date <= query_end_date:
                if current_quote.close and prev_quote.close and prev_quote.close > 0:
                    five_day_change = ((current_quote.close - prev_quote.close) / prev_quote.close) * 100
                    five_day_change = round(five_day_change, 2)
                    
                    # 更新数据库，不管之前是否有值都更新
                    update_query = f"""
                        UPDATE {table_name} 
                        SET five_day_change_percent = :five_day_change
                        WHERE code = :stock_code AND date = :date
                    """
                    
                    db.execute(text(update_query), {
                        "five_day_change": five_day_change,
                        "stock_code": request.stock_code,
                        "date": current_quote.date
                    })
                    
                    updated_count += 1
        
        # 提交事务
        db.commit()
        
        message = f"股票 {request.stock_code} 在 {start_date_fmt} 到 {end_date_fmt} 期间的5天升跌%计算完成"
        print(f"[calculate_five_day_change] {message}, 更新了 {updated_count} 条记录 (数据源: {'港股' if is_hk_data else 'A股'})")
        
        return {
            "message": message,
            "stock_code": request.stock_code,
            "start_date": start_date_fmt,
            "end_date": end_date_fmt,
            "updated_count": updated_count,
            "total_records": len(quotes)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        error_msg = f"计算5天升跌%失败: {str(e)}"
        print(f"[calculate_five_day_change] {error_msg}")
        raise HTTPException(status_code=500, detail=error_msg)

@router.post("/calculate_ten_day_change")
def calculate_ten_day_change(
    request: CalculateFiveDayChangeRequest,
    db: Session = Depends(get_db)
):
    """计算指定日期范围内股票的10天涨跌%（支持A股和港股）"""
    try:
        # 格式化日期
        start_date_fmt = format_date_yyyymmdd(request.start_date)
        end_date_fmt = format_date_yyyymmdd(request.end_date)
        
        if not start_date_fmt or not end_date_fmt:
            raise HTTPException(status_code=400, detail="日期格式无效")
        
        # 如果提供了扩展后的结束日期，使用它来扩展查询和更新范围
        query_end_date = format_date_yyyymmdd(request.extended_end_date) if request.extended_end_date else end_date_fmt
        
        print(f"[calculate_ten_day_change] 开始计算股票 {request.stock_code} 在 {start_date_fmt} 到 {end_date_fmt} 期间的10天涨跌%（查询范围扩展到 {query_end_date}）")
        
        # 为了确保最后10条记录也能计算10天涨跌%，需要查询更早的数据
        # 计算开始日期前10个工作日的数据
        extended_start_date = _get_date_before_business_days(start_date_fmt, 10)
        
        # 1. 先尝试从A股表查询
        extended_query_a = """
            SELECT date, close
            FROM historical_quotes 
            WHERE code = :stock_code 
            AND date >= :extended_start_date 
            AND date <= :end_date
            ORDER BY date ASC
        """
        
        result_a = db.execute(text(extended_query_a), {
            "stock_code": request.stock_code,
            "extended_start_date": extended_start_date,
            "end_date": query_end_date
        })
        
        quotes = result_a.fetchall()
        is_hk_data = False
        
        # 2. 如果A股表完全没有数据，才从港股表查询
        if len(quotes) == 0:
            print(f"[calculate_ten_day_change] A股表无数据，从港股表查询: code={request.stock_code}")
            extended_query_hk = """
                SELECT date, close
                FROM historical_quotes_hk 
                WHERE code = :stock_code 
                AND date >= :extended_start_date 
                AND date <= :end_date
                ORDER BY date ASC
            """
            
            result_hk = db.execute(text(extended_query_hk), {
                "stock_code": request.stock_code,
                "extended_start_date": extended_start_date,
                "end_date": query_end_date
            })
            
            quotes = result_hk.fetchall()
            is_hk_data = True
        
        if len(quotes) < 11:
            raise HTTPException(status_code=400, detail="数据不足11天，无法计算10天涨跌%")
        
        # 计算10天涨跌%
        updated_count = 0
        table_name = "historical_quotes_hk" if is_hk_data else "historical_quotes"
        
        for i in range(10, len(quotes)):
            current_quote = quotes[i]
            prev_quote = quotes[i-10]  # 10天前的数据
            
            # 更新到扩展结束日期范围内的所有记录，确保最后10天的数据也能被计算
            if current_quote.date >= start_date_fmt and current_quote.date <= query_end_date:
                if current_quote.close and prev_quote.close and prev_quote.close > 0:
                    ten_day_change = ((current_quote.close - prev_quote.close) / prev_quote.close) * 100
                    ten_day_change = round(ten_day_change, 2)
                    
                    # 更新数据库，不管之前是否有值都更新
                    update_query = f"""
                        UPDATE {table_name} 
                        SET ten_day_change_percent = :ten_day_change
                        WHERE code = :stock_code AND date = :date
                    """
                    
                    db.execute(text(update_query), {
                        "ten_day_change": ten_day_change,
                        "stock_code": request.stock_code,
                        "date": current_quote.date
                    })
                    
                    updated_count += 1
        
        # 提交事务
        db.commit()
        
        message = f"股票 {request.stock_code} 在 {start_date_fmt} 到 {end_date_fmt} 期间的10天涨跌%计算完成"
        print(f"[calculate_ten_day_change] {message}, 更新了 {updated_count} 条记录 (数据源: {'港股' if is_hk_data else 'A股'})")
        
        return {
            "message": message,
            "stock_code": request.stock_code,
            "start_date": start_date_fmt,
            "end_date": end_date_fmt,
            "updated_count": updated_count,
            "total_records": len(quotes)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        error_msg = f"计算10天涨跌%失败: {str(e)}"
        print(f"[calculate_ten_day_change] {error_msg}")
        raise HTTPException(status_code=500, detail=error_msg)

@router.post("/calculate_thirty_day_change")
def calculate_thirty_day_change(
    request: CalculateFiveDayChangeRequest,
    db: Session = Depends(get_db)
):
    """计算指定日期范围内股票的30天涨跌%（支持A股和港股）"""
    try:
        start_date_fmt = format_date_yyyymmdd(request.start_date)
        end_date_fmt = format_date_yyyymmdd(request.end_date)

        if not start_date_fmt or not end_date_fmt:
            raise HTTPException(status_code=400, detail="日期格式无效")

        print(f"[calculate_thirty_day_change] 开始计算股票 {request.stock_code} 在 {start_date_fmt} 到 {end_date_fmt} 期间的30天涨跌%")

        extended_start_date = _get_date_before_business_days(start_date_fmt, 30)
        query_end_date = format_date_yyyymmdd(request.extended_end_date) if request.extended_end_date else end_date_fmt

        # 1. 先尝试从A股表查询
        extended_query_a = """
            SELECT date, close
            FROM historical_quotes 
            WHERE code = :stock_code 
            AND date >= :extended_start_date 
            AND date <= :end_date
            ORDER BY date ASC
        """

        result_a = db.execute(text(extended_query_a), {
            "stock_code": request.stock_code,
            "extended_start_date": extended_start_date,
            "end_date": query_end_date
        })

        quotes = result_a.fetchall()
        is_hk_data = False

        # 2. 如果A股表完全没有数据，才从港股表查询
        if len(quotes) == 0:
            print(f"[calculate_thirty_day_change] A股表无数据，从港股表查询: code={request.stock_code}")
            extended_query_hk = """
                SELECT date, close
                FROM historical_quotes_hk 
                WHERE code = :stock_code 
                AND date >= :extended_start_date 
                AND date <= :end_date
                ORDER BY date ASC
            """

            result_hk = db.execute(text(extended_query_hk), {
                "stock_code": request.stock_code,
                "extended_start_date": extended_start_date,
                "end_date": query_end_date
            })

            quotes = result_hk.fetchall()
            is_hk_data = True

        if len(quotes) < 31:
            raise HTTPException(status_code=400, detail="数据不足31天，无法计算30天涨跌%")

        updated_count = 0
        table_name = "historical_quotes_hk" if is_hk_data else "historical_quotes"
        
        for i in range(30, len(quotes)):
            current_quote = quotes[i]
            prev_quote = quotes[i - 30]

            # 更新到扩展结束日期范围内的所有记录，确保最后30天的数据也能被计算
            if current_quote.date >= start_date_fmt and current_quote.date <= query_end_date:
                if current_quote.close and prev_quote.close and prev_quote.close > 0:
                    thirty_day_change = ((current_quote.close - prev_quote.close) / prev_quote.close) * 100
                    thirty_day_change = round(thirty_day_change, 2)

                    update_query = f"""
                        UPDATE {table_name} 
                        SET thirty_day_change_percent = :thirty_day_change
                        WHERE code = :stock_code AND date = :date
                    """

                    db.execute(text(update_query), {
                        "thirty_day_change": thirty_day_change,
                        "stock_code": request.stock_code,
                        "date": current_quote.date
                    })

                    updated_count += 1

        db.commit()

        message = f"股票 {request.stock_code} 在 {start_date_fmt} 到 {end_date_fmt} 期间的30天涨跌%计算完成"
        print(f"[calculate_thirty_day_change] {message}, 更新了 {updated_count} 条记录 (数据源: {'港股' if is_hk_data else 'A股'})")

        return {
            "message": message,
            "stock_code": request.stock_code,
            "start_date": start_date_fmt,
            "end_date": end_date_fmt,
            "updated_count": updated_count,
            "total_records": len(quotes)
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        error_msg = f"计算30天涨跌%失败: {str(e)}"
        print(f"[calculate_thirty_day_change] {error_msg}")
        raise HTTPException(status_code=500, detail=error_msg)

@router.post("/calculate_sixty_day_change")
def calculate_sixty_day_change(
    request: CalculateFiveDayChangeRequest,
    db: Session = Depends(get_db)
):
    """计算指定日期范围内股票的60天涨跌%（支持A股和港股）"""
    try:
        # 格式化日期
        start_date_fmt = format_date_yyyymmdd(request.start_date)
        end_date_fmt = format_date_yyyymmdd(request.end_date)
        
        # 如果提供了扩展后的结束日期，使用它来扩展查询范围
        # 但只更新原始日期范围内的记录
        query_end_date = format_date_yyyymmdd(request.extended_end_date) if request.extended_end_date else end_date_fmt
        
        if not start_date_fmt or not end_date_fmt:
            raise HTTPException(status_code=400, detail="日期格式无效")
        
        print(f"[calculate_sixty_day_change] 开始计算股票 {request.stock_code} 在 {start_date_fmt} 到 {end_date_fmt} 期间的60天涨跌%（查询范围扩展到 {query_end_date}）")
        
        # 为了确保最后60条记录也能计算60天涨跌%，需要查询更早的数据
        # 计算开始日期前60个工作日的数据
        extended_start_date = _get_date_before_business_days(start_date_fmt, 60)
        
        # 1. 先尝试从A股表查询
        extended_query_a = """
            SELECT date, close
            FROM historical_quotes 
            WHERE code = :stock_code 
            AND date >= :extended_start_date 
            AND date <= :end_date
            ORDER BY date ASC
        """
        
        result_a = db.execute(text(extended_query_a), {
            "stock_code": request.stock_code,
            "extended_start_date": extended_start_date,
            "end_date": query_end_date  # 使用扩展后的结束日期进行查询
        })
        
        quotes = result_a.fetchall()
        is_hk_data = False
        
        # 2. 如果A股表完全没有数据，才从港股表查询
        if len(quotes) == 0:
            print(f"[calculate_sixty_day_change] A股表无数据，从港股表查询: code={request.stock_code}")
            extended_query_hk = """
                SELECT date, close
                FROM historical_quotes_hk 
                WHERE code = :stock_code 
                AND date >= :extended_start_date 
                AND date <= :end_date
                ORDER BY date ASC
            """
            
            result_hk = db.execute(text(extended_query_hk), {
                "stock_code": request.stock_code,
                "extended_start_date": extended_start_date,
                "end_date": query_end_date
            })
            
            quotes = result_hk.fetchall()
            is_hk_data = True
        
        if len(quotes) < 61:
            raise HTTPException(status_code=400, detail="数据不足61天，无法计算60天涨跌%")
        
        print(f"[calculate_sixty_day_change] 查询到 {len(quotes)} 条记录，日期范围: {quotes[0].date if quotes else 'N/A'} 到 {quotes[-1].date if quotes else 'N/A'}")
        
        # 计算60天涨跌%
        updated_count = 0
        table_name = "historical_quotes_hk" if is_hk_data else "historical_quotes"
        
        for i in range(60, len(quotes)):
            current_quote = quotes[i]
            prev_quote = quotes[i-60]  # 60天前的数据
            
            # 更新所有能计算的记录，不限制日期范围
            # 因为查询时已经扩展了范围（从extended_start_date到query_end_date），
            # 所以查询结果中的所有记录都应该被更新（只要它们有足够的前60天数据）
            # 这样可以确保最后一条记录也能被计算
            if current_quote.close and prev_quote.close and prev_quote.close > 0:
                sixty_day_change = ((current_quote.close - prev_quote.close) / prev_quote.close) * 100
                sixty_day_change = round(sixty_day_change, 2)
                
                # 更新数据库，不管之前是否有值都更新
                update_query = f"""
                    UPDATE {table_name} 
                    SET sixty_day_change_percent = :sixty_day_change
                    WHERE code = :stock_code AND date = :date
                """
                
                db.execute(text(update_query), {
                    "sixty_day_change": sixty_day_change,
                    "stock_code": request.stock_code,
                    "date": current_quote.date
                })
                
                updated_count += 1
                print(f"[calculate_sixty_day_change] 更新记录: date={current_quote.date}, sixty_day_change={sixty_day_change}%")
        
        # 提交事务
        db.commit()
        
        message = f"股票 {request.stock_code} 在 {start_date_fmt} 到 {end_date_fmt} 期间的60天涨跌%计算完成"
        print(f"[calculate_sixty_day_change] {message}, 更新了 {updated_count} 条记录 (数据源: {'港股' if is_hk_data else 'A股'})")
        
        return {
            "message": message,
            "stock_code": request.stock_code,
            "start_date": start_date_fmt,
            "end_date": end_date_fmt,
            "updated_count": updated_count,
            "total_records": len(quotes)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        error_msg = f"计算60天涨跌%失败: {str(e)}"
        print(f"[calculate_sixty_day_change] {error_msg}")
        raise HTTPException(status_code=500, detail=error_msg)