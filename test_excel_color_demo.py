#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
测试Excel导出功能的颜色效果演示
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import io

def demo_excel_export_colors():
    """演示Excel导出颜色效果"""
    
    print("📊 演示Excel导出颜色效果...")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "历史行情数据颜色演示"
    
    # 定义颜色样式
    red_font = Font(color="FF0000", bold=True)    # 上涨红色
    green_font = Font(color="00AA00", bold=True)  # 下跌绿色
    header_font = Font(bold=True)
    
    # 表头
    headers = [
        "股票代码", "股票名称", "日期", "开盘", "收盘", "最高", "最低",
        "成交量(万手)", "成交额(亿)", "涨跌幅%", "涨跌额", "换手率%",
        "5天涨跌%", "10天涨跌%", "60天涨跌%"
    ]
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
    
    # 模拟数据（来自您的图片）
    sample_data = [
        ("300223", "北京君正", "2025-09-19", 81.01, 81.70, 36.57, 30.40, 1.30, 1.05, 8.70, 14.07, 3.55, 24.49),
        ("300223", "北京君正", "2025-09-18", 82.24, 80.65, 39.28, 32.42, -1.92, -1.58, 9.34, 9.13, 7.28, 25.04),
        ("300223", "北京君正", "2025-09-17", 83.00, 82.23, 26.55, 21.95, -0.75, -0.62, 6.31, 15.30, 6.50, 25.35),
        ("300223", "北京君正", "2025-09-16", 83.00, 82.85, 31.48, 26.13, -0.90, -0.75, 7.48, 10.39, 9.14, 25.53),
        ("300223", "北京君正", "2025-09-15", 88.50, 83.60, 52.34, 43.99, -1.59, -1.35, 12.44, 8.00, 8.30, 27.44),
    ]
    
    # 写入数据并应用颜色格式
    for row_idx, row in enumerate(sample_data, 2):
        # 基本数据
        ws.cell(row=row_idx, column=1, value=row[0])  # 股票代码
        ws.cell(row=row_idx, column=2, value=row[1])  # 股票名称
        ws.cell(row=row_idx, column=3, value=row[2])  # 日期
        ws.cell(row=row_idx, column=4, value=row[3])  # 开盘
        
        # 收盘价（需要颜色格式）
        close_cell = ws.cell(row=row_idx, column=5, value=row[4])
        change_percent = row[7]  # 涨跌幅
        if change_percent > 0:
            close_cell.font = red_font
        elif change_percent < 0:
            close_cell.font = green_font
        
        ws.cell(row=row_idx, column=6, value=row[5])  # 最高
        ws.cell(row=row_idx, column=7, value=row[6])  # 最低
        ws.cell(row=row_idx, column=8, value=f"{row[5]:.2f}")  # 成交量(万手)
        ws.cell(row=row_idx, column=9, value=f"{row[6]:.2f}")  # 成交额(亿)
        
        # 涨跌幅%（需要颜色格式）
        change_pct_cell = ws.cell(row=row_idx, column=10, value=f"{change_percent:.2f}%")
        if change_percent > 0:
            change_pct_cell.font = red_font
        elif change_percent < 0:
            change_pct_cell.font = green_font
        
        # 涨跌额（需要颜色格式）
        change_cell = ws.cell(row=row_idx, column=11, value=row[9])
        if change_percent > 0:
            change_cell.font = red_font
        elif change_percent < 0:
            change_cell.font = green_font
        
        ws.cell(row=row_idx, column=12, value=f"{row[5]:.2f}%")  # 换手率%
        
        # 各期涨跌%（需要颜色格式）
        for col_offset, pct_val in enumerate([row[10], row[11], row[12]], 13):
            pct_cell = ws.cell(row=row_idx, column=col_offset, value=f"{pct_val:.2f}%")
            if pct_val > 0:
                pct_cell.font = red_font
            elif pct_val < 0:
                pct_cell.font = green_font
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 保存文件
    filename = "历史行情颜色效果演示.xlsx"
    wb.save(filename)
    
    print(f"✅ Excel文件已生成: {filename}")
    print("📋 颜色效果说明:")
    print("   🔴 红色字体: 表示上涨数据（涨跌幅 > 0）")
    print("   🟢 绿色字体: 表示下跌数据（涨跌幅 < 0）")
    print("   ⚫ 黑色字体: 表示其他数据")
    print("\n🎯 应用范围:")
    print("   • 收盘价")
    print("   • 涨跌幅%")
    print("   • 涨跌额")
    print("   • 5天/10天/60天涨跌%")

if __name__ == "__main__":
    demo_excel_export_colors()